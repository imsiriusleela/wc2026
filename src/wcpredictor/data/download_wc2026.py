"""Fetch and parse the WC2026 tournament sheet from football-data.co.uk.

football-data.co.uk publishes a 'WorldCup2026' sheet in WorldCup_fdco.xlsx
a few days before kickoff.  Run this script once the sheet is live:

    uv run python -m wcpredictor.data.download_wc2026

Outputs
-------
data/raw/wc2026_fixtures.csv
    Columns: date, team_a, team_b, neutral, goals_a, goals_b
    goals_a / goals_b are empty strings for unplayed matches.

Side-effects
------------
Prints the fresh SHA-256 of WorldCup_fdco.xlsx so config.FDCO_ODDS_SHA256
can be updated if the file changed.  The WorldCup2026 sheet is picked up
automatically by load_wc_odds() once it exists in the xlsx.
"""

from __future__ import annotations

import hashlib
import io
import ssl
import sys
import urllib.request
from pathlib import Path

import openpyxl
import pandas as pd

from wcpredictor.config import DATA_RAW, FDCO_ODDS_URL
from wcpredictor.data.normalize_teams import canonical

_DEST = DATA_RAW / "WorldCup_fdco.xlsx"
_FIXTURES_OUT = DATA_RAW / "wc2026_fixtures.csv"
_WC2026_SHEET = "WorldCup2026"


def _sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _fetch_xlsx() -> bytes:
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    print(f"Fetching {FDCO_ODDS_URL} …")
    with urllib.request.urlopen(FDCO_ODDS_URL, context=ctx) as resp:
        return resp.read()


def _parse_wc2026_sheet(xlsx_bytes: bytes) -> pd.DataFrame:
    """Return DataFrame with: date, team_a, team_b, neutral, goals_a, goals_b.

    goals_a / goals_b are None for future/unplayed matches.
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    if _WC2026_SHEET not in wb.sheetnames:
        wb.close()
        return pd.DataFrame()

    ws = wb[_WC2026_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return pd.DataFrame()

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]

    def _col(name: str) -> int:
        return headers.index(name)

    # Column names mirror WorldCup2022 layout
    home_col = _col("Home")
    away_col = _col("Away")
    date_col = _col("Date")

    # Goals: HGFT / AGFT (full-time 90-min goals); absent or None if unplayed
    hg_col = headers.index("HGFT") if "HGFT" in headers else None
    ag_col = headers.index("AGFT") if "AGFT" in headers else None

    records: list[dict] = []
    for row in rows[1:]:
        home_raw = row[home_col]
        away_raw = row[away_col]
        date_raw = row[date_col]
        if home_raw is None or away_raw is None or date_raw is None:
            continue

        ta = canonical(str(home_raw).strip())
        tb = canonical(str(away_raw).strip())
        if not ta or not tb:
            continue

        date = pd.Timestamp(date_raw)

        goals_a: int | None = None
        goals_b: int | None = None
        if hg_col is not None and ag_col is not None:
            try:
                goals_a = int(row[hg_col])
                goals_b = int(row[ag_col])
            except (TypeError, ValueError):
                goals_a = None
                goals_b = None

        records.append({
            "date": date.strftime("%Y-%m-%d"),
            "team_a": ta,
            "team_b": tb,
            "neutral": True,  # all WC matches played at neutral sites
            "goals_a": "" if goals_a is None else goals_a,
            "goals_b": "" if goals_b is None else goals_b,
        })

    return pd.DataFrame(records)


def download_wc2026(force_refetch: bool = False) -> Path | None:
    """Fetch WorldCup_fdco.xlsx and parse WC2026 fixtures.

    Returns path to wc2026_fixtures.csv if successful, None if the sheet
    is not yet available.
    """
    DATA_RAW.mkdir(parents=True, exist_ok=True)

    if _DEST.exists() and not force_refetch:
        xlsx_bytes = _DEST.read_bytes()
        current_sha = _sha256(xlsx_bytes)
        # Check if the WC2026 sheet is already present in the cached file
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
        has_sheet = _WC2026_SHEET in wb.sheetnames
        wb.close()
        if has_sheet:
            print(f"Using cached xlsx (SHA-256: {current_sha})")
        else:
            # Re-fetch to see if fdco has published the sheet since last download
            xlsx_bytes = _fetch_xlsx()
            new_sha = _sha256(xlsx_bytes)
            if new_sha != current_sha:
                print(f"xlsx updated. New SHA-256: {new_sha}")
                print("Update FDCO_ODDS_SHA256 in config.py with the new hash.")
                _DEST.write_bytes(xlsx_bytes)
            else:
                print(f"SHA-256 unchanged: {current_sha}")
    else:
        xlsx_bytes = _fetch_xlsx()
        fresh_sha = _sha256(xlsx_bytes)
        print(f"Fresh SHA-256: {fresh_sha}")
        print("If this differs from config.FDCO_ODDS_SHA256, update config.py.")
        _DEST.write_bytes(xlsx_bytes)

    df = _parse_wc2026_sheet(xlsx_bytes)

    if df.empty:
        print(
            f"\n'{_WC2026_SHEET}' sheet not yet in the xlsx "
            f"(fdco publishes it ~2 days before kickoff).\n"
            "Retry closer to 2026-06-11 kickoff."
        )
        return None

    _FIXTURES_OUT.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(_FIXTURES_OUT, index=False)
    print(f"Wrote {len(df)} fixtures → {_FIXTURES_OUT}")
    return _FIXTURES_OUT


if __name__ == "__main__":
    force = "--force" in sys.argv
    result = download_wc2026(force_refetch=force)
    if result is None:
        sys.exit(1)
