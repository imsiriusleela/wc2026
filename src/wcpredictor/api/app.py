"""FastAPI application for the World Cup predictor.

Endpoints:
  GET  /health             - liveness check
  GET  /teams              - sorted list of known teams
  GET  /predict            - live single-match prediction (cached frozen state)
  GET  /fixtures           - precomputed fixture predictions
  GET  /tournament         - precomputed tournament simulation
  POST /refresh-odds       - pull fresh fdco odds file and clear prediction cache
  GET  /                   - static frontend (index.html)
"""
from __future__ import annotations

import ast
import glob
import hashlib
import json
import math
import os
import threading
from contextlib import asynccontextmanager
from pathlib import Path
from typing import Any, Literal

import pandas as pd
from fastapi import FastAPI, HTTPException, Query
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles

from wcpredictor.config import DATA_PROCESSED, DATA_RAW, FDCO_ODDS_SHA256, TOURNAMENT_START
from wcpredictor.data.download_odds import download_odds
from wcpredictor.data.normalize_teams import canonical
from wcpredictor.features.odds import load_wc_odds
from wcpredictor.predict import _build_frozen_state, _predict_one_frozen

from .schemas import (
    FixtureRow, PredictResponse, RefreshOddsResponse, RefreshResultsResponse,
    ResimulateResponse, ScorecardResponse, TeamStanding, TournamentResponse,
)

STATIC_DIR = Path(__file__).parent / "static"


def _default_as_of() -> str:
    """Return today's date, but never earlier than TOURNAMENT_START."""
    today = pd.Timestamp.today().strftime("%Y-%m-%d")
    return max(TOURNAMENT_START, today)

# Module-level cache: (as_of, frozenset(models)) -> frozen state dict
_STATE_CACHE: dict[tuple[str, frozenset], dict] = {}

# Guards concurrent calls to POST /refresh-odds
_REFRESH_LOCK = threading.Lock()


def _get_state(as_of: str, models: list[str]) -> dict:
    key = (as_of, frozenset(models))
    if key not in _STATE_CACHE:
        _STATE_CACHE[key] = _build_frozen_state(as_of, models)
    return _STATE_CACHE[key]


def _latest_artifact(pattern: str) -> Path | None:
    matches = sorted(glob.glob(str(DATA_PROCESSED / pattern)))
    return Path(matches[-1]) if matches else None


def _known_teams() -> list[str]:
    ratings_path = DATA_PROCESSED / "elo_ratings.csv"
    if ratings_path.exists():
        df = pd.read_csv(ratings_path)
        col = "team" if "team" in df.columns else df.columns[0]
        return sorted(df[col].dropna().unique().tolist())

    fixtures_path = DATA_RAW / "wc2026_fixtures.csv"
    if fixtures_path.exists():
        df = pd.read_csv(fixtures_path)
        teams = set(df["team_a"].tolist()) | set(df["team_b"].tolist())
        return sorted(canonical(t) for t in teams)

    pred_path = _latest_artifact("wc2026_predictions_*.csv")
    if pred_path:
        df = pd.read_csv(pred_path)
        teams = set(df["team_a"].tolist()) | set(df["team_b"].tolist())
        return sorted(teams)

    return []


def _parse_matrix(raw: Any) -> list[list[float]] | None:
    if raw is None:
        return None
    if isinstance(raw, list):
        return raw
    try:
        parsed = ast.literal_eval(str(raw))
        return parsed
    except Exception:
        return None


def _parse_scorelines(raw: Any) -> list[dict[str, Any]] | None:
    if raw is None:
        return None
    if isinstance(raw, list):
        return raw
    try:
        parsed = ast.literal_eval(str(raw))
        return parsed
    except Exception:
        return None


@asynccontextmanager
async def _lifespan(app: FastAPI):
    try:
        _get_state(_default_as_of(), ["poisson", "dixon_coles", "ensemble"])
    except Exception:
        pass
    yield


app = FastAPI(title="WC2026 Predictor", version="0.1.0", lifespan=_lifespan)


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/teams", response_model=list[str])
def teams() -> list[str]:
    return _known_teams()


@app.get("/predict", response_model=PredictResponse)
def predict(
    team_a: str = Query(..., description="First team name"),
    team_b: str = Query(..., description="Second team name"),
    model: Literal["poisson", "dixon_coles", "ensemble", "ensemble_mkt"] = Query(
        "ensemble_mkt", description="Model to use"
    ),
    neutral: bool = Query(True, description="Neutral venue"),
    as_of: str = Query(None, description="Data cutoff date (ISO); defaults to today or TOURNAMENT_START"),
) -> PredictResponse:
    if as_of is None:
        as_of = _default_as_of()
    team_a_c = canonical(team_a)
    team_b_c = canonical(team_b)

    known = set(_known_teams())
    if known and team_a_c not in known:
        raise HTTPException(
            status_code=422,
            detail=f"Unknown team: {team_a!r}. Use /teams to see valid options.",
        )
    if known and team_b_c not in known:
        raise HTTPException(
            status_code=422,
            detail=f"Unknown team: {team_b!r}. Use /teams to see valid options.",
        )

    models_needed = (
        ["poisson"] if model == "poisson"
        else ["poisson", "dixon_coles"] if model == "dixon_coles"
        else ["poisson", "dixon_coles", "ensemble"]
        if model in {"ensemble", "ensemble_mkt"}
        else ["poisson"]
    )

    try:
        state = _get_state(as_of, models_needed)
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"Model state unavailable: {exc}")

    try:
        result = _predict_one_frozen(state, model, team_a_c, team_b_c, neutral)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))

    matrix = result.get("score_matrix")
    if isinstance(matrix, list) and matrix and not isinstance(matrix[0], list):
        matrix = None

    return PredictResponse(
        team_a=team_a_c,
        team_b=team_b_c,
        neutral=neutral,
        model=model,
        model_version=result["model_version"],
        p_win=result["p_win"],
        p_draw=result["p_draw"],
        p_loss=result["p_loss"],
        lambda_a=result.get("lambda_a", math.nan),
        lambda_b=result.get("lambda_b", math.nan),
        elo_a=result["elo_a"],
        elo_b=result["elo_b"],
        score_matrix=matrix or [],
        top_scorelines=result.get("top_scorelines") or [],
        markets=result.get("markets"),
    )


@app.get("/fixtures", response_model=list[FixtureRow])
def fixtures(
    model: str | None = Query(None, description="Filter by model name"),
) -> list[FixtureRow]:
    path = _latest_artifact("wc2026_predictions_*.csv")
    if path is None:
        raise HTTPException(status_code=503, detail="No precomputed fixture predictions found.")

    df = pd.read_csv(path)
    if model is not None:
        df = df[df["model"] == model]

    rows: list[FixtureRow] = []
    for _, row in df.iterrows():
        rows.append(
            FixtureRow(
                team_a=str(row.get("team_a", "")),
                team_b=str(row.get("team_b", "")),
                date=str(row.get("date", "")),
                neutral=bool(row.get("neutral", True)),
                model=str(row.get("model", "")),
                model_version=str(row.get("model_version", "")),
                p_win=float(row["p_win"]) if pd.notna(row.get("p_win")) else None,
                p_draw=float(row["p_draw"]) if pd.notna(row.get("p_draw")) else None,
                p_loss=float(row["p_loss"]) if pd.notna(row.get("p_loss")) else None,
                lambda_a=float(row["lambda_a"]) if pd.notna(row.get("lambda_a")) else None,
                lambda_b=float(row["lambda_b"]) if pd.notna(row.get("lambda_b")) else None,
                elo_a=float(row["elo_a"]) if pd.notna(row.get("elo_a")) else None,
                elo_b=float(row["elo_b"]) if pd.notna(row.get("elo_b")) else None,
                score_matrix=_parse_matrix(row.get("score_matrix")),
                top_scorelines=_parse_scorelines(row.get("top_scorelines")),
            )
        )
    return rows


@app.get("/tournament", response_model=TournamentResponse)
def tournament() -> TournamentResponse:
    csv_path = _latest_artifact("wc2026_tournament_sim_*.csv")
    json_path = _latest_artifact("wc2026_tournament_sim_*.json")

    if csv_path is None:
        raise HTTPException(status_code=503, detail="No precomputed tournament simulation found.")

    df = pd.read_csv(csv_path)
    standings = [
        TeamStanding(
            team=str(r.team),
            group=str(r.group),
            p_win_group=float(r.p_win_group),
            p_runner_up=float(r.p_runner_up),
            p_r32=float(r.p_r32),
            p_r16=float(r.p_r16),
            p_qf=float(r.p_qf),
            p_sf=float(r.p_sf),
            p_final=float(r.p_final),
            p_champion=float(r.p_champion),
        )
        for _, r in df.iterrows()
    ]

    meta: dict[str, Any] = {}
    top10: list[dict[str, Any]] = []
    if json_path is not None:
        try:
            meta = json.loads(json_path.read_text())
            top10 = meta.get("top10_champion", [])
        except Exception:
            pass

    if not top10:
        top10 = (
            df.sort_values("p_champion", ascending=False)
            .head(10)[["team", "group", "p_champion"]]
            .to_dict(orient="records")
        )

    return TournamentResponse(
        as_of=str(meta.get("as_of", TOURNAMENT_START)),
        model=str(meta.get("model", "poisson")),
        n_sims=int(meta.get("n_sims", 500)),
        n_group_fixed=int(meta.get("n_group_fixed", 0)),
        n_ko_fixed=int(meta.get("n_ko_fixed", 0)),
        standings=standings,
        top10_champion=top10,
    )


@app.get("/scorecard", response_model=ScorecardResponse)
def scorecard() -> ScorecardResponse:
    path = DATA_PROCESSED / "wc2026_scorecard.json"
    if not path.exists():
        raise HTTPException(status_code=503, detail="Scorecard not yet available.")
    data = json.loads(path.read_text())
    return ScorecardResponse(**data)


@app.post("/refresh-odds", response_model=RefreshOddsResponse)
def refresh_odds() -> RefreshOddsResponse:
    """Pull fresh odds from fdco and/or the-odds-api (h2h + AH/O-U).

    Requires an explicit user action (button click) — honors the no-silent-scraping rule.
    Clears _STATE_CACHE so the next /predict rebuilds with updated odds.
    One call refreshes both 1X2 (used by ensemble_mkt blend) and AH/O-U (matrix blend)
    since both read the same odds_api_wc2026.json.
    Does NOT auto-rewrite config.py; the new fdco SHA is surfaced for manual re-pinning.
    Note: Fixtures and Tournament tabs use pre-generated snapshots and are not affected.
    """
    acquired = _REFRESH_LOCK.acquire(blocking=False)
    if not acquired:
        raise HTTPException(status_code=409, detail="Refresh already in progress.")
    try:
        from wcpredictor.data.download_odds_api import fetch_odds_api

        note_parts: list[str] = []
        errors: list[str] = []

        # fdco xlsx — non-fatal
        actual_sha = ""
        sha_changed = False
        fdco_ok = False
        try:
            download_odds(force=True, verify=False)
            fdco_ok = True
        except Exception as exc:
            errors.append(f"fdco: {exc}")

        odds_path = DATA_RAW / "WorldCup_fdco.xlsx"
        if odds_path.exists():
            actual_sha = hashlib.sha256(odds_path.read_bytes()).hexdigest()
            sha_changed = actual_sha != FDCO_ODDS_SHA256

        # the-odds-api — non-fatal; also refreshes AH/O-U (same JSON)
        api_key = os.environ.get("ODDS_API_KEY", "").strip()
        odds_api_refreshed = False
        if api_key:
            try:
                fetch_odds_api(force=True)
                odds_api_refreshed = True
            except Exception as exc:
                errors.append(f"odds-api: {exc}")
        else:
            note_parts.append("ODDS_API_KEY not set — skipped live API refresh.")

        # Count 2026 rows from merged odds (fdco + live JSON)
        try:
            odds_df = load_wc_odds()
            n_2026 = int((odds_df["year"] == 2026).sum()) if "year" in odds_df.columns else 0
        except Exception:
            n_2026 = 0

        # Breakdown by source
        n_fdco = 0
        n_api = 0
        try:
            from wcpredictor.features.odds import load_wc_odds as _lwo_fdco
            import openpyxl as _opx
            wb = _opx.load_workbook(odds_path, read_only=True, data_only=True)
            if "WorldCup2026" in wb.sheetnames:
                ws = wb["WorldCup2026"]
                n_fdco = max(0, sum(1 for _ in ws.iter_rows(min_row=2)) - 0)
            wb.close()
        except Exception:
            pass
        n_api = max(0, n_2026 - n_fdco)

        _STATE_CACHE.clear()

        if n_2026 > 0:
            note_parts.append(f"{n_2026} WC 2026 matches priced (fdco={n_fdco}, api={n_api}) — next prediction uses market odds.")
        else:
            note_parts.append("No WC 2026 odds available yet.")
        if sha_changed:
            note_parts.append(
                f"fdco file changed (new SHA: {actual_sha[:16]}…). "
                "Re-pin FDCO_ODDS_SHA256 in config.py for reproducibility."
            )
        if errors:
            note_parts.append("Errors: " + "; ".join(errors))
        note_parts.append(
            "Fixtures and Tournament tabs use pre-generated snapshots; "
            "re-run predict_fixtures / simulate CLI to update them."
        )

        if errors and n_2026 == 0:
            raise HTTPException(status_code=502, detail=" ".join(note_parts))

        return RefreshOddsResponse(
            status="ok",
            n_odds_2026=n_2026,
            n_odds_2026_fdco=n_fdco,
            n_odds_2026_api=n_api,
            odds_api_refreshed=odds_api_refreshed,
            file_sha256=actual_sha,
            pinned_sha256=FDCO_ODDS_SHA256,
            sha_changed=sha_changed,
            state_cache_cleared=True,
            note=" ".join(note_parts),
        )
    finally:
        _REFRESH_LOCK.release()


@app.post("/refresh-results", response_model=RefreshResultsResponse)
def refresh_results() -> RefreshResultsResponse:
    """Pull latest martj42 results, update the WC2026 results store, and mark played fixtures.

    Writes to data/raw/results_master.csv and data/raw/wc2026_results.csv (never touches pinned results.csv).
    Clears _STATE_CACHE so the next /predict uses rolled-forward Elo ratings.
    Non-fatal on network error.
    """
    acquired = _REFRESH_LOCK.acquire(blocking=False)
    if not acquired:
        raise HTTPException(status_code=409, detail="Refresh already in progress.")
    try:
        from wcpredictor.data.results_2026 import update_wc2026_results, mark_fixtures_played
        from wcpredictor.evaluation.live import run_refresh

        errors: list[str] = []
        stats: dict = {"n_total": 0, "n_new": 0, "n_group": 0, "n_knockout": 0}
        n_fixtures_updated = 0

        try:
            stats = update_wc2026_results()
        except Exception as exc:
            errors.append(f"results: {exc}")

        try:
            n_fixtures_updated = mark_fixtures_played()
        except Exception as exc:
            errors.append(f"fixtures: {exc}")

        _STATE_CACHE.clear()

        try:
            run_refresh(as_of_date=_default_as_of())
        except Exception as exc:
            errors.append(f"scorecard: {exc}")

        note_parts = [
            f"{stats['n_total']} results in store ({stats['n_new']} new, "
            f"{stats['n_group']} group, {stats['n_knockout']} knockout).",
            f"{n_fixtures_updated} fixture(s) marked as played.",
            "Prediction cache cleared — next /predict uses updated ratings.",
        ]
        if errors:
            note_parts.append("Errors: " + "; ".join(errors))

        return RefreshResultsResponse(
            status="ok",
            n_results_total=stats["n_total"],
            n_new=stats["n_new"],
            n_group=stats["n_group"],
            n_knockout=stats["n_knockout"],
            n_fixtures_updated=n_fixtures_updated,
            note=" ".join(note_parts),
        )
    finally:
        _REFRESH_LOCK.release()


@app.post("/resimulate", response_model=ResimulateResponse)
def resimulate(
    n_sims: int = Query(5000, description="Number of Monte-Carlo simulations"),
    model: str = Query("ensemble_mkt", description="Model to use for simulation"),
) -> ResimulateResponse:
    """Rerun tournament simulation conditioned on played results.

    Synchronous; typically takes 1-2 minutes.  Writes dated artifacts to
    data/processed/; /tournament picks them up via _latest_artifact.
    """
    acquired = _REFRESH_LOCK.acquire(blocking=False)
    if not acquired:
        raise HTTPException(status_code=409, detail="Refresh already in progress.")
    try:
        from wcpredictor.simulate import simulate_tournament

        as_of = _default_as_of()
        try:
            df = simulate_tournament(
                as_of=as_of,
                model=model,
                n_sims=n_sims,
                condition_on_results=True,
            )
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"Simulation failed: {exc}")

        # Read meta from the JSON artifact just written
        sim_json = DATA_PROCESSED / f"wc2026_tournament_sim_{as_of}.json"
        meta: dict = {}
        if sim_json.exists():
            try:
                meta = json.loads(sim_json.read_text())
            except Exception:
                pass

        out_csv = str(DATA_PROCESSED / f"wc2026_tournament_sim_{as_of}.csv")
        return ResimulateResponse(
            status="ok",
            as_of=as_of,
            model=model,
            n_sims=n_sims,
            n_group_fixed=int(meta.get("n_group_fixed", 0)),
            n_ko_fixed=int(meta.get("n_ko_fixed", 0)),
            output_csv=out_csv,
            note=f"Simulation complete. Conditioned on {meta.get('n_group_fixed', 0)} group + "
                 f"{meta.get('n_ko_fixed', 0)} KO results.",
        )
    finally:
        _REFRESH_LOCK.release()


app.mount("/", StaticFiles(directory=str(STATIC_DIR), html=True), name="static")
