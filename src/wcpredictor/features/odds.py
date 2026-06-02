"""Parse football-data.co.uk WC xlsx → margin-normalised implied probabilities.

The xlsx (WorldCup_fdco.xlsx) contains sheets WorldCup2014 / 2018 / 2022, each
with 64 rows (one per match) and columns including H-Avg, D-Avg, A-Avg (the
market-average decimal odds across bookmakers).

We remove the bookmaker margin by normalising:
    p_i = (1/o_i) / sum(1/o_j for j in {H,D,A})

so that p_win + p_draw + p_loss = 1.

Home/Away in the xlsx corresponds to team_a/team_b in our matches dataset.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd

from wcpredictor.config import DATA_RAW
from wcpredictor.data.normalize_teams import canonical

_SHEET_YEARS: dict[str, int] = {
    "WorldCup2014": 2014,
    "WorldCup2018": 2018,
    "WorldCup2022": 2022,
    "WorldCup2026": 2026,  # sheet appears on fdco a few days before kickoff; skipped if absent
}

# All three average-market columns must be present
_AVG_COLS = ("H-Avg", "D-Avg", "A-Avg")


def _implied_probs(h: float, d: float, a: float) -> tuple[float, float, float]:
    p_h, p_d, p_a = 1.0 / h, 1.0 / d, 1.0 / a
    total = p_h + p_d + p_a
    return p_h / total, p_d / total, p_a / total


def load_wc_odds(xlsx_path: Path | None = None) -> pd.DataFrame:
    """Return DataFrame with columns:
        year, date, team_a, team_b, p_win, p_draw, p_loss

    p_win  = implied prob that team_a wins in 90 min.
    p_loss = implied prob that team_b wins in 90 min.

    Covers WC 2010 (betexplorer CSV), 2014, 2018, 2022 (football-data.co.uk xlsx).
    """
    if xlsx_path is None:
        xlsx_path = DATA_RAW / "WorldCup_fdco.xlsx"

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    records: list[dict] = []

    for sheet_name, year in _SHEET_YEARS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h) if h is not None else "" for h in rows[0]]

        # Locate the average-odds columns; skip sheet if missing
        try:
            h_col = headers.index("H-Avg")
            d_col = headers.index("D-Avg")
            a_col = headers.index("A-Avg")
        except ValueError:
            continue

        home_col = headers.index("Home")
        away_col = headers.index("Away")
        date_col = headers.index("Date")

        for row in rows[1:]:
            try:
                h_odds = float(row[h_col])
                d_odds = float(row[d_col])
                a_odds = float(row[a_col])
            except (TypeError, ValueError):
                continue
            if h_odds <= 1.0 or d_odds <= 1.0 or a_odds <= 1.0:
                continue

            raw_date = row[date_col]
            if raw_date is None:
                continue
            date = pd.Timestamp(raw_date)

            ta = canonical(str(row[home_col]))
            tb = canonical(str(row[away_col]))
            if not ta or not tb:
                continue

            p_win, p_draw, p_loss = _implied_probs(h_odds, d_odds, a_odds)
            records.append(
                dict(
                    year=year,
                    date=date,
                    team_a=ta,
                    team_b=tb,
                    p_win=p_win,
                    p_draw=p_draw,
                    p_loss=p_loss,
                )
            )

    wb.close()
    df_xlsx = pd.DataFrame(records)

    # Merge WC2010 odds from betexplorer CSV (if available)
    wc2010_csv = DATA_RAW / "wc2010_odds.csv"
    if wc2010_csv.exists():
        df2010 = pd.read_csv(wc2010_csv, parse_dates=["date"])
        df2010["team_a"] = df2010["home"].apply(canonical)
        df2010["team_b"] = df2010["away"].apply(canonical)
        df2010[["p_win", "p_draw", "p_loss"]] = df2010.apply(
            lambda r: pd.Series(_implied_probs(r["odds_h"], r["odds_d"], r["odds_a"])),
            axis=1,
        )
        df2010["year"] = 2010
        df2010 = df2010[["year", "date", "team_a", "team_b", "p_win", "p_draw", "p_loss"]]
        df_xlsx = pd.concat([df2010, df_xlsx], ignore_index=True)

    return df_xlsx


def merge_odds_features(
    features_df: pd.DataFrame,
    odds_df: pd.DataFrame,
) -> pd.DataFrame:
    """Attach odds_p_win / odds_p_draw / odds_p_loss / has_odds to features_df.

    Keying: (year, team_a, team_b) extracted from the date column.
    Symmetric: if (a, b) is in odds_df, (b, a) is also registered with W/L swapped.
    Rows with no odds match get NaN for the three probability columns and 0.0 for
    has_odds.  NaN is intentional — HGB handles it natively; zero-fill would lie.

    Parameters
    ----------
    features_df : must contain date, team_a, team_b columns.
    odds_df     : output of load_wc_odds() — year, team_a, team_b, p_win, p_draw, p_loss.

    Returns
    -------
    Copy of features_df with four new columns appended.
    """
    lookup: dict[tuple[int, str, str], tuple[float, float, float]] = {}
    for _, r in odds_df.iterrows():
        ta, tb = str(r.team_a), str(r.team_b)
        yr = int(r.year)
        fwd = (float(r.p_win), float(r.p_draw), float(r.p_loss))
        lookup[(yr, ta, tb)] = fwd
        lookup[(yr, tb, ta)] = (fwd[2], fwd[1], fwd[0])  # W/L swap

    years = features_df["date"].dt.year.to_numpy(int)
    ta_arr = features_df["team_a"].astype(str).to_numpy()
    tb_arr = features_df["team_b"].astype(str).to_numpy()

    p_wins: list[float] = []
    p_draws: list[float] = []
    p_losses: list[float] = []
    has_odds_vals: list[float] = []

    for yr, ta, tb in zip(years, ta_arr, tb_arr):
        key = (int(yr), ta, tb)
        if key in lookup:
            pw, pd_, pl = lookup[key]
            p_wins.append(pw)
            p_draws.append(pd_)
            p_losses.append(pl)
            has_odds_vals.append(1.0)
        else:
            p_wins.append(float("nan"))
            p_draws.append(float("nan"))
            p_losses.append(float("nan"))
            has_odds_vals.append(0.0)

    out = features_df.copy()
    out["odds_p_win"] = p_wins
    out["odds_p_draw"] = p_draws
    out["odds_p_loss"] = p_losses
    out["has_odds"] = has_odds_vals
    return out


def align_odds_to_test(
    odds_df: pd.DataFrame,
    year: int,
    test_elo: pd.DataFrame,
) -> list[list[float]] | None:
    """Align odds rows to test_elo by (team_a, team_b) within the given WC year.

    Lookup is symmetric: if (a, b) is in the odds with [p_win, p_draw, p_loss],
    then (b, a) is also registered as [p_loss, p_draw, p_win] (W/L swap).
    This handles betexplorer home/away ordering differing from our dataset.

    Returns a list of [p_win, p_draw, p_loss] in the same row order as test_elo,
    or None if fewer than 50 % of test rows can be matched (signals unusable data).
    Unmatched rows fall back to uniform [1/3, 1/3, 1/3].
    """
    yr_odds = odds_df[odds_df["year"] == year]
    if yr_odds.empty:
        return None

    lookup: dict[tuple[str, str], list[float]] = {}
    for _, r in yr_odds.iterrows():
        ta, tb = str(r.team_a), str(r.team_b)
        probs = [r.p_win, r.p_draw, r.p_loss]
        lookup[(ta, tb)] = probs
        # Symmetric: reversed fixture swaps win/loss
        lookup[(tb, ta)] = [r.p_loss, r.p_draw, r.p_win]

    result: list[list[float]] = []
    matched = 0
    for _, row in test_elo.iterrows():
        key = (str(row.team_a), str(row.team_b))
        if key in lookup:
            result.append(lookup[key])
            matched += 1
        else:
            result.append([1 / 3, 1 / 3, 1 / 3])

    if matched < len(result) * 0.5:
        return None
    return result
